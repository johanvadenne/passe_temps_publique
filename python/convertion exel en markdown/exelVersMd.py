import openpyxl

table = [[]]
x = 0
texte = ""

document = openpyxl.load_workbook(r"chemin")
for sheet_name in document.sheetnames:
    print(sheet_name)

# Sélectionnez une feuille de calcul spécifique
sheet = document['nom_de_la_feuille']

# Afficher le contenu de toutes les cellules dans la feuille de calcul
for row in sheet.iter_rows():
    
    for cell in row:
        table[x].append(cell.value)
    x+=1
    table.append([])

for i in table:
    texte += "\n"
    for j in i:
        if j == None:
            j = ""
        texte += " | " + str(j)
    texte += " |"

print(texte)

document = open("fichier.md", "w+", encoding="utf8")
document.write(texte)
document.close